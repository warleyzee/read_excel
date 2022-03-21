from datetime import date, datetime
from tkinter.font import BOLD
from turtle import back
from webbrowser import BackgroundBrowser
import pandas as pd
import matplotlib
import os
import itertools

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, PatternFill
from posixpath import sep
from read_save_pdf import Read_Save_PDF
from move_file import MoveFile
from pprint import pprint
from datetime import datetime, date

class File_Excel():

    def save_file_exce(self,):
        font = Font(name=" Microsoft YaHei ", bold=True)
        alignment = Alignment(vertical="top", wrap_text=True)
        pattern_fill = PatternFill(fill_type="solid", fgColor="D7E4BC")
        side = Side(style="thin")
        border = Border(left=side, right=side, top=side, bottom=side)

        #variavel que recebe os dados para preencher na tabela
        dados = Read_Save_PDF().create_list_pdf()
        

        df = pd.read_excel(r"C:\Users\Warley Souza\Music\read_excel\Cube_finish.xlsx", engine='openpyxl')
        # df.head()
        # df["Date Cast"] = df["Date Cast"].dt.strftime("%d-%m-%Y")
        # print(df['Date Cast'])
        # input()
        survey_df = pd.DataFrame(df)

        for i in dados:
            cube = survey_df['Cube'] == i['Client Ref']
            survey_df.loc[cube,'Date Tested'] = i['Date of Test']
            survey_df.loc[cube,'Density'] = i['Density']
            if i['Test Age'] == '7':
                survey_df.loc[cube,'7 Day Result']  = i['Compressive Strength']
            elif i['Test Age'] == '14':
                survey_df.loc[cube,'14 Day Result'] = i['Compressive Strength']
            elif i['Test Age'] == '28':
                survey_df.loc[cube,'28 Day Result'] = i['Compressive Strength']
            elif i['Test Age'] == '56':
                survey_df.loc[cube,'56 Day Result'] = i['Compressive Strength']
            else:
                survey_df.loc[cube,'7 Day Result']  = i['Compressive Strength']+ '(' + i['Test Age'] + 'Days' + ')'
                pass
            survey_df
        try:
            print()
            os.remove(r'C:\Users\Warley Souza\Music\read_excel\Cube_finish.xlsx')
            file_name = ('Cube_finish.xlsx')
            writer = pd.ExcelWriter('Cube_finish.xlsx',
                        engine='openpyxl',
                        date_format='dd mmmm yyyy')
            survey_df.to_excel(writer, sheet_name='Sheet1', index=False)
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            border = Border(left=side, right=side, top=side, bottom=side)
            for cell in itertools.chain(*worksheet["A1:J1"]):
                cell.font = font
                cell.alignment = alignment
                cell.fill = pattern_fill
                cell.border = border
            for cell in itertools.chain(*worksheet["A2:J600"]):
                cell.border = border
            for cell in itertools.chain(*worksheet["C20:C60"], *worksheet["E20:E60"]):
                cell.number_format = '#,##0.00'
            for cell in itertools.chain(*worksheet["D20:D60"]):
                cell.number_format = '0%'
            worksheet.column_dimensions["A"].width = 8
            worksheet.column_dimensions["B"].width = 21
            worksheet.column_dimensions["C"].width = 35
            worksheet.column_dimensions["D"].width = 19
            worksheet.column_dimensions["E"].width = 9
            worksheet.column_dimensions["F"].width = 14
            worksheet.column_dimensions["G"].width = 15
            worksheet.column_dimensions["H"].width = 15
            worksheet.column_dimensions["I"].width = 15
            worksheet.column_dimensions["j"].width = 18
            writer.save()
            workbook

            # move_xlsx = MoveFile().move_file_xlsx()
            print('Data Frame is written to Excel File Sucessfully')
        except:
            print()
            print("Error")

test = File_Excel()
test.save_file_exce()          
